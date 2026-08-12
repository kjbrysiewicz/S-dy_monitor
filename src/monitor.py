"""Monitor orzecznictwa dot. środków unijnych — aktualizacja mastera.

Dopisuje nowe rekordy do arkusza "Baza orzeczeń" w pliku
"Baza_orzecznictwa_po_stronie_beneficjenta___MASTER" zgodnie z jego konwencjami:
- ID w schemacie ORZ-NNNN (kontynuacja numeracji),
- zasada jednego rekordu: deduplikacja po znormalizowanej sygnaturze
  oraz po "Kluczu sprawy" (PL-{sygnatura}-{rok}),
- wartości słownikowe: Status publikacji = "Do weryfikacji",
  Status weryfikacji = "Metadane wstępne", Wynik = "Do oceny",
  Etap sporu = "Inne / do klasyfikacji" (do ręcznej klasyfikacji),
- pola merytoryczne (teza publiczna, znaczenie, rozstrzygnięcie, podstawa
  prawna) pozostają PUSTE — zgodnie z Instrukcją uzupełnia się je po lekturze
  pełnej treści orzeczenia,
- style i formaty liczb kopiowane z istniejącego wiersza danych,
- rozszerzanie zakresu autofiltra i walidacji danych na nowe wiersze.

Przebieg: pobranie pliku z Drive -> zapytania SAOS (opcjonalnie CBOSA)
-> deduplikacja -> [opcjonalnie: pobranie pełnych tekstów i analiza słownikowa]
-> dopisanie -> przeliczenie formuł (LibreOffice, jeśli dostępne) -> wysyłka.

Tryby:
  python src/monitor.py                # standard: okno lookback_days
  python src/monitor.py --backfill     # jednorazowo: od daty backfill.since
                                       # z pełną paginacją i analizą treści
"""

from __future__ import annotations

import argparse
import copy
import logging
import re
import shutil
import subprocess
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path

import yaml
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

import analyze  # noqa: E402
import drive_sync  # noqa: E402
import saos_client  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
log = logging.getLogger("monitor")

ROOT = Path(__file__).parent.parent
LOCAL_XLSX = ROOT / "master.xlsx"

# Kolumny mastera (arkusz "Baza orzeczeń") — kolejność i nazwy muszą się zgadzać
EXPECTED_COLUMNS = [
    "ID", "Sygnatura", "Rok", "Sąd", "Jurysdykcja", "Rodzaj finansowania",
    "Etap sporu", "Problemy prawne", "Wynik dla beneficjenta", "Data orzeczenia",
    "Rozstrzygnięcie", "Podstawa prawna", "Teza publiczna",
    "Znaczenie dla beneficjenta", "Powiązane orzeczenia – publiczne", "Priorytet",
    "Status publikacji", "Ryzyko publikacji", "Status weryfikacji", "Typ źródła",
    "Tytuł reprezentatywny", "Link źródłowy", "Liczba kopii", "Kategorie źródłowe",
    "Notatka wewnętrzna", "Ścieżki źródłowe", "Wszystkie URL",
    "Data modyfikacji źródła", "Typy plików", "Redaktor", "Data weryfikacji",
    "Data publikacji", "Slug / adres", "Tytuł SEO", "Uwagi końcowe", "Klucz sprawy",
]

JURYSDYKCJA_BY_COURT_TYPE = {
    "ADMINISTRATIVE": "Sądowoadministracyjna",
    "COMMON": "Cywilna / powszechna",
    "SUPREME": "Cywilna / powszechna",
}


def load_config() -> dict:
    with open(ROOT / "config.yaml", encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def norm_sig(s: str) -> str:
    """Normalizacja sygnatury do porównań: wielkie litery, pojedyncze spacje."""
    return re.sub(r"\s+", " ", str(s)).strip().upper()


def cand_key(item: dict) -> str:
    """Klucz deduplikacji kandydata.

    Sygnatury sądów powszechnych powtarzają się między sądami (wiele
    "I C 100/23" w kraju), więc dla jurysdykcji cywilnej klucz obejmuje
    także sąd. Dla spraw sądowoadministracyjnych i SN sygnatura jest
    unikalna w skali kraju.
    """
    base = norm_sig(item["sygnatura"])
    if item.get("jurysdykcja") == "Cywilna / powszechna" and "Najwyższy" not in str(item.get("sad", "")):
        return base + "|" + norm_sig(item.get("sad", ""))
    return base


def year_from_signature(sygnatura: str, fallback: str) -> int | None:
    """Rok = końcówka sygnatury (II GSK 1128/15 -> 2015); awaryjnie rok orzeczenia."""
    m = re.search(r"/(\d{2,4})\s*$", sygnatura)
    if m:
        y = int(m.group(1))
        if y < 100:
            y += 2000 if y <= 49 else 1900
        return y
    if fallback and len(fallback) >= 4 and fallback[:4].isdigit():
        return int(fallback[:4])
    return None


def short_court_name(name: str) -> str:
    n = (name or "").strip()
    if "Naczelny Sąd Administracyjny" in n:
        return "NSA"
    m = re.search(r"Wojewódzki Sąd Administracyjny\s*(w|we)\s+(.+)", n)
    if m:
        return f"WSA {m.group(1)} {m.group(2)}".strip()
    return n or "NSA/WSA"


def last_data_row(ws) -> int:
    last = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value not in (None, ""):
            last = r
    return last


def next_orz_id(ws, last_row: int) -> int:
    best = 0
    for r in range(2, last_row + 1):
        v = str(ws.cell(r, 1).value or "")
        m = re.match(r"ORZ-(\d+)", v)
        if m:
            best = max(best, int(m.group(1)))
    return best + 1


def extend_ranges(ws, new_max_row: int) -> None:
    """Rozszerza autofiltr i zakresy walidacji danych na nowe wiersze."""
    try:
        ws.auto_filter.ref = f"A1:AJ{new_max_row}"
    except Exception:  # noqa: BLE001
        pass
    try:
        for dv in ws.data_validations.dataValidation:
            for rng in dv.sqref.ranges:
                if rng.max_row < new_max_row:
                    rng.max_row = new_max_row
    except Exception as exc:  # noqa: BLE001
        log.warning("Nie udało się rozszerzyć walidacji danych: %s", exc)


def try_recalc(path: Path) -> None:
    """Przelicza formuły (Dashboard) LibreOffice'em, jeśli jest dostępny.

    openpyxl nie zapisuje wyników formuł — bez przeliczenia podgląd pliku na
    Drive pokazywałby puste komórki Dashboardu (Excel/Arkusze przeliczą je
    same przy otwarciu, więc brak LibreOffice nie jest krytyczny).
    """
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        log.info("LibreOffice niedostępny — pomijam przeliczenie formuł.")
        return
    try:
        outdir = path.parent / "_recalc"
        outdir.mkdir(exist_ok=True)
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", str(outdir), str(path)],
            check=True, capture_output=True, timeout=120,
        )
        produced = outdir / path.name
        if produced.exists():
            shutil.move(str(produced), str(path))
            log.info("Formuły przeliczone.")
    except Exception as exc:  # noqa: BLE001
        log.warning("Przeliczenie formuł nie powiodło się (%s) — kontynuuję.", exc)


def parse_judgment_date(value: str) -> datetime | None:
    try:
        return datetime.strptime(value[:10], "%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def collect_candidates(cfg: dict, date_from: date, per_kw: int) -> dict[str, dict]:
    """Zwraca kandydatów {znormalizowana sygnatura: rekord} ze wszystkich źródeł."""
    candidates: dict[str, dict] = {}
    excerpt_len = int(cfg["excerpt_length"])

    if cfg.get("saos", {}).get("enabled", True):
        for group in cfg["keyword_groups"]:
            court_types = group.get("court_types", ["ADMINISTRATIVE"])
            for kw in group["keywords"]:
                phrase = kw["phrase"]
                for item in saos_client.search_keyword(
                    phrase, date_from, court_types,
                    max_results=per_kw, excerpt_length=excerpt_len,
                ):
                    item["finansowanie"] = group["finansowanie"]
                    item["problem"] = kw.get("problem", "")
                    item["jurysdykcja"] = JURYSDYKCJA_BY_COURT_TYPE.get(
                        item.get("court_type", "ADMINISTRATIVE"),
                        "Sądowoadministracyjna",
                    )
                    key = cand_key(item)
                    if key in candidates:
                        continue
                    candidates[key] = item

    cbosa_cfg = cfg.get("cbosa", {})
    if cbosa_cfg.get("enabled", False):
        try:
            import cbosa_client

            delay = float(cbosa_cfg.get("delay_seconds", 3))
            symbols = [str(s) for s in cbosa_cfg.get("symbols", [])]

            # Filtr wstępny po kategorii sprawy: wyszukiwanie po symbolu
            # (np. 6559 — środki unijne) łapie WSZYSTKIE sprawy z kategorii,
            # niezależnie od sformułowań w treści.
            if cbosa_cfg.get("search_by_symbol", True):
                for symbol in symbols:
                    for item in cbosa_client.search_by_symbol(
                        symbol, date_from,
                        delay_seconds=delay, excerpt_length=excerpt_len,
                    ):
                        item["finansowanie"] = "Fundusze UE / polityka spójności"
                        item["problem"] = ""
                        item["jurysdykcja"] = "Sądowoadministracyjna"
                        item["symbol_ok"] = True  # trafione po symbolu
                        key = cand_key(item)
                        if key in candidates:
                            continue
                        candidates[key] = item

            # Wyszukiwanie po frazach, zawężone do symbolu (jeśli podano)
            if cbosa_cfg.get("search_by_keywords", False):
                narrowing = symbols[0] if symbols else None
                for group in cfg["keyword_groups"]:
                    if "ADMINISTRATIVE" not in group.get("court_types", ["ADMINISTRATIVE"]):
                        continue
                    for kw in group["keywords"]:
                        for item in cbosa_client.search_keyword(
                            kw["phrase"], date_from,
                            delay_seconds=delay, excerpt_length=excerpt_len,
                            symbol=narrowing,
                        ):
                            item["finansowanie"] = group["finansowanie"]
                            item["problem"] = kw.get("problem", "")
                            item["jurysdykcja"] = "Sądowoadministracyjna"
                            if narrowing:
                                item["symbol_ok"] = True
                            key = cand_key(item)
                            if key in candidates:
                                continue
                            candidates[key] = item
        except Exception as exc:  # noqa: BLE001
            log.error("CBOSA: moduł zawiódł, kontynuuję bez niego: %s", exc)

    # --- Źródła cywilne: Portal Orzeczeń Sądów Powszechnych i baza SN ---
    civil_groups = [
        g for g in cfg["keyword_groups"]
        if {"COMMON", "SUPREME"} & set(g.get("court_types", []))
    ]

    def _add_civil(item: dict, group: dict, kw: dict) -> None:
        item["finansowanie"] = group["finansowanie"]
        item["problem"] = kw.get("problem", "")
        item["jurysdykcja"] = "Cywilna / powszechna"
        key = cand_key(item)
        if key not in candidates:
            candidates[key] = item

    if cfg.get("ms_portal", {}).get("enabled", False):
        try:
            import ms_client

            delay = float(cfg["ms_portal"].get("delay_seconds", 3))
            for group in civil_groups:
                for kw in group["keywords"]:
                    for item in ms_client.search_keyword(
                        kw["phrase"], date_from,
                        delay_seconds=delay, excerpt_length=excerpt_len,
                    ):
                        _add_civil(item, group, kw)
        except Exception as exc:  # noqa: BLE001
            log.error("PO SP: moduł zawiódł, kontynuuję bez niego: %s", exc)

    if cfg.get("sn_portal", {}).get("enabled", False):
        try:
            import sn_client

            delay = float(cfg["sn_portal"].get("delay_seconds", 3))
            for group in civil_groups:
                for kw in group["keywords"]:
                    for item in sn_client.search_keyword(
                        kw["phrase"], date_from,
                        delay_seconds=delay, excerpt_length=excerpt_len,
                    ):
                        _add_civil(item, group, kw)
        except Exception as exc:  # noqa: BLE001
            log.error("SN: moduł zawiódł, kontynuuję bez niego: %s", exc)

    return candidates


def build_row(item: dict, orz_num: int, defaults: dict) -> dict[str, object]:
    """Buduje słownik {nazwa kolumny: wartość} dla nowego rekordu."""
    sygnatura = re.sub(r"\s+", " ", item["sygnatura"]).strip()
    rok = year_from_signature(sygnatura, item.get("data", ""))
    dt = parse_judgment_date(item.get("data", ""))
    zrodlo = item.get("zrodlo", "SAOS")
    typ_zrodla = {
        "CBOSA": "CBOSA – pełny tekst publiczny",
        "SAOS": "SAOS – indeks publiczny",
        "PO SP": "Portal Orzeczeń SP – pełny tekst publiczny",
        "SN": "Baza orzeczeń SN – pełny tekst publiczny",
    }.get(zrodlo, zrodlo)
    today = datetime.now().strftime("%Y-%m-%d")
    fragment = (item.get("fragment") or "").strip()
    notatka = f"[AUTO {zrodlo} {today}] trafienie: „{item['fraza']}”"
    if item.get("symbol_info"):
        notatka += f"; {item['symbol_info']}"
    if item.get("analiza_info"):
        notatka += f"; analiza treści: {item['analiza_info']}"
    if fragment:
        notatka += f"; początek treści: {fragment}"

    return {
        "ID": f"ORZ-{orz_num:04d}",
        "Sygnatura": sygnatura,
        "Rok": rok,
        "Sąd": short_court_name(item.get("sad", "")),
        "Jurysdykcja": item.get("jurysdykcja", "Sądowoadministracyjna"),
        "Rodzaj finansowania": item.get("finansowanie", ""),
        "Etap sporu": item.get("etap_auto") or defaults["etap_sporu"],
        "Problemy prawne": item.get("problemy_auto") or item.get("problem", ""),
        "Wynik dla beneficjenta": defaults["wynik"],
        "Data orzeczenia": dt,
        "Status publikacji": defaults["status_publikacji"],
        "Status weryfikacji": defaults["status_weryfikacji"],
        "Typ źródła": typ_zrodla,
        "Tytuł reprezentatywny": f"{sygnatura} — {item['fraza']}",
        "Link źródłowy": item.get("link", ""),
        "Kategorie źródłowe": item.get("fraza", ""),
        "Notatka wewnętrzna": notatka,
        "Klucz sprawy": _case_key(sygnatura, rok, item),
    }


def _case_key(sygnatura: str, rok: int | None, item: dict) -> str:
    """Klucz sprawy wg konwencji mastera: PL-{sygnatura}-{rok}.

    Sygnatury sądów powszechnych powtarzają się między sądami, więc dla
    jurysdykcji cywilnej (poza SN) klucz jest rozszerzany o sąd — inaczej
    dwie różne sprawy dostałyby identyczny klucz.
    """
    key = f"PL-{sygnatura}-{rok}" if rok else f"PL-{sygnatura}"
    sad = str(item.get("sad", ""))
    if item.get("jurysdykcja") == "Cywilna / powszechna" and "Najwyższy" not in sad:
        skrot = re.sub(r"\s+", " ", sad).strip()
        if skrot:
            key += f"-{skrot}"
    return key


def filter_by_symbol(items: list[dict], cfg: dict) -> list[dict]:
    """Filtr wstępny po kategorii sprawy (symbol CBOSA, np. 6559).

    Dotyczy wyłącznie spraw sądowoadministracyjnych. Kandydaci z CBOSA
    trafieni po symbolu przechodzą bez ponownej weryfikacji. Kandydatom
    z SAOS symbol jest sprawdzany po sygnaturze na stronie CBOSA.
    Sprawy cywilne (PFR) nie mają symboli CBOSA — przechodzą bez filtra.
    """
    sf = cfg.get("symbol_filter", {})
    if not sf.get("enabled", False):
        return items
    symbols = {str(s) for s in sf.get("symbols", ["6559"])}
    keep_unverified = bool(sf.get("keep_unverified", True))
    delay = float(sf.get("delay_seconds", 1.5))

    import cbosa_client

    kept: list[dict] = []
    checked = 0
    for item in items:
        if item.get("jurysdykcja") != "Sądowoadministracyjna" or item.get("symbol_ok"):
            kept.append(item)
            continue
        found = cbosa_client.get_symbols(item["sygnatura"], delay_seconds=delay)
        time.sleep(delay)
        checked += 1
        if found is None:
            if keep_unverified:
                item["symbol_info"] = "symbol nieustalony"
                kept.append(item)
            else:
                log.info("Symbol nieustalony — odrzucam %s", item["sygnatura"])
            continue
        if found & symbols:
            item["symbol_ok"] = True
            item["symbol_info"] = "symbol " + "/".join(sorted(found & symbols))
            kept.append(item)
        else:
            log.info("Poza kategorią (symbole %s) — odrzucam %s",
                     ",".join(sorted(found)), item["sygnatura"])
        if checked % 25 == 0:
            log.info("Weryfikacja symboli: sprawdzono %d", checked)
    log.info("Filtr symboli: %d/%d kandydatów zachowanych", len(kept), len(items))
    return kept


def analyze_items(items: list[dict]) -> None:
    """Pobiera pełne teksty z SAOS i klasyfikuje (Problemy prawne, Etap sporu)."""
    total = len(items)
    for i, item in enumerate(items, 1):
        if item.get("zrodlo") != "SAOS" or not item.get("saos_id"):
            continue
        text = saos_client.get_full_text(item["saos_id"])
        time.sleep(0.4)
        if not text:
            continue
        res = analyze.classify(text, base_problem=item.get("problem", ""))
        item["problemy_auto"] = res["problemy"]
        item["etap_auto"] = res["etap"]
        if res["trafienia"]:
            top = "; ".join(f"{k} ({v})" for k, v in list(res["trafienia"].items())[:4])
            item["analiza_info"] = f"klasyfikacja słownikowa wg trafień: {top}"
        if i % 25 == 0 or i == total:
            log.info("Analiza treści: %d/%d", i, total)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--backfill", action="store_true",
                        help="pobierz wszystkie orzeczenia od backfill.since i analizuj treść")
    args = parser.parse_args()

    cfg = load_config()
    sheet_name = cfg["master"]["sheet_name"]
    defaults = cfg["master"]["defaults"]

    bf = cfg.get("backfill", {})
    if args.backfill:
        date_from = date.fromisoformat(str(bf.get("since", "2023-01-01")))
        per_kw = int(bf.get("max_results_per_keyword", 1000))
        do_analysis = bool(bf.get("fetch_full_text", True))
        log.info("TRYB BACKFILL: od %s, do %d wyników/frazę, analiza treści: %s",
                 date_from, per_kw, do_analysis)
    else:
        date_from = date.today() - timedelta(days=int(cfg["lookback_days"]))
        per_kw = int(cfg["max_results_per_keyword"])
        do_analysis = bool(cfg.get("analyze_in_standard_mode", True))

    # 1. Pobierz master z Drive
    drive_sync.download_xlsx(str(LOCAL_XLSX))
    wb = load_workbook(LOCAL_XLSX)
    if sheet_name not in wb.sheetnames:
        log.error("Brak arkusza '%s' w pliku.", sheet_name)
        return 1
    ws = wb[sheet_name]

    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    if header[: len(EXPECTED_COLUMNS)] != EXPECTED_COLUMNS:
        diff = [
            f"kol.{i+1}: jest '{h}', oczekiwano '{e}'"
            for i, (h, e) in enumerate(zip(header, EXPECTED_COLUMNS))
            if h != e
        ]
        log.error("Struktura mastera odbiega od oczekiwanej: %s", "; ".join(diff[:5]))
        return 1
    col_idx = {name: i + 1 for i, name in enumerate(EXPECTED_COLUMNS)}

    last_row = last_data_row(ws)
    known: set[str] = set()
    for r in range(2, last_row + 1):
        sig = ws.cell(r, col_idx["Sygnatura"]).value
        sad = ws.cell(r, col_idx["Sąd"]).value
        jur = str(ws.cell(r, col_idx["Jurysdykcja"]).value or "")
        if sig:
            if "cywiln" in jur.lower() and "najwyższy" not in str(sad or "").lower():
                known.add(norm_sig(sig) + "|" + norm_sig(sad or ""))
            else:
                known.add(norm_sig(sig))
        klucz = ws.cell(r, col_idx["Klucz sprawy"]).value
        if klucz:
            known.add(norm_sig(klucz))
    log.info(
        "Master: %d rekordów (ostatni wiersz %d), okno wyszukiwania od %s",
        last_row - 1, last_row, date_from,
    )

    # 2. Zbierz kandydatów
    candidates = collect_candidates(cfg, date_from, per_kw)
    new_items = [
        v for k, v in candidates.items()
        if k not in known
        and norm_sig(v["sygnatura"]) not in known
        and norm_sig(f"PL-{v['sygnatura']}") not in known
    ]
    new_items.sort(key=lambda x: x.get("data") or "")
    log.info("Znaleziono %d orzeczeń, w tym nowych: %d", len(candidates), len(new_items))

    if not new_items:
        log.info("Brak nowych orzeczeń — plik bez zmian, pomijam upload.")
        return 0

    # 3. Filtr wstępny po kategorii sprawy (symbol CBOSA)
    new_items = filter_by_symbol(new_items, cfg)
    if not new_items:
        log.info("Po filtrze symboli brak nowych orzeczeń — pomijam upload.")
        return 0

    # 3a. Analiza pełnych tekstów (klasyfikacja słownikowa)
    if do_analysis:
        analyze_items(new_items)

    # 4. Dopisz rekordy (style i formaty kopiowane z ostatniego wiersza danych)
    orz_num = next_orz_id(ws, last_row)
    template_row = last_row if last_row >= 2 else 2
    row_idx = last_row

    for item in new_items:
        row_idx += 1
        values = build_row(item, orz_num, defaults)
        orz_num += 1
        for name, c in col_idx.items():
            src = ws.cell(template_row, c)
            dst = ws.cell(row_idx, c, value=values.get(name))
            dst.font = copy.copy(src.font)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format
            if name == "Data orzeczenia" and values.get(name):
                dst.number_format = "yyyy-MM-dd"
            if name == "Link źródłowy" and values.get(name):
                dst.hyperlink = values[name]

    extend_ranges(ws, row_idx)
    wb.save(LOCAL_XLSX)

    # 5. Przelicz formuły Dashboardu i odeślij na Drive
    try_recalc(LOCAL_XLSX)
    drive_sync.upload_xlsx(str(LOCAL_XLSX))
    log.info(
        "Dopisano %d nowych rekordów (ID do ORZ-%04d).", len(new_items), orz_num - 1
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
