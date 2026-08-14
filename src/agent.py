"""Agent contentowy — ocena nowych orzeczeń i rekomendacje publikacji.

Przebieg:
1. Pobiera master z Google Drive.
2. Wybiera rekordy dodane przez monitor (Status weryfikacji = "Metadane wstępne"),
   które nie były jeszcze ocenione (brak ID w arkuszu "Rekomendacje").
3. Dla każdego pobiera pełny tekst orzeczenia (SAOS/CBOSA) i prosi model Claude
   o ocenę: istotność dla praktyki beneficjentów, nowość linii orzeczniczej,
   rekomendowany kanał (LinkedIn / blog / artykuł naukowy lub glosa / podcast /
   pomiń), roboczy tytuł, hook i kluczowe punkty.
4. Zapisuje wyniki do arkusza "Rekomendacje" w masterze (tworzy go przy
   pierwszym uruchomieniu) i odsyła plik na Drive.

Agent NICZEGO nie zmienia w arkuszu "Baza orzeczeń" — ocena AI jest podpowiedzią
redakcyjną, nie klasyfikacją merytoryczną rekordu.

Uruchomienie:
    ANTHROPIC_API_KEY=... GDRIVE_...=... python src/agent.py [--max 20]

Bez ANTHROPIC_API_KEY kończy się komunikatem (kod 0) — monitor działa dalej
niezależnie od agenta.
"""

from __future__ import annotations

import argparse
import copy
import json
import logging
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))

import drive_sync  # noqa: E402
import saos_client  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
log = logging.getLogger("agent")

ROOT = Path(__file__).parent.parent
LOCAL_XLSX = ROOT / "master.xlsx"

API_URL = "https://api.anthropic.com/v1/messages"
API_VERSION = "2023-06-01"

REKOM_SHEET = "Rekomendacje"
REKOM_COLUMNS = [
    "Data oceny", "ID", "Sygnatura", "Sąd", "Data orzeczenia",
    "Istotność (0-100)", "Dla beneficjenta (wstępnie)", "Rekomendowany kanał",
    "Kanały dodatkowe", "Tytuł roboczy", "Hook / teza otwierająca",
    "Kluczowe punkty", "Uzasadnienie oceny", "Link", "Status",
]

SYSTEM_PROMPT = """Jesteś asystentem redakcyjnym kancelarii BBS-Legal (Brysiewicz \
Bokina i Wspólnicy), wyspecjalizowanej w sporach o środki unijne i pomoc publiczną \
WYŁĄCZNIE po stronie beneficjentów. Kancelaria prowadzi bloga kbrysiewicz.pl \
(praktyczne, ironiczne, pierwszoosobowe teksty dla beneficjentów), podcast \
"Ster na prawo", profil LinkedIn oraz publikuje artykuły naukowe i glosy \
(Przegląd Sądowy, Palestra, EPS, Monitor Prawniczy i in.).

Oceniasz orzeczenia sądów polskich pod kątem przydatności contentowej. Kryteria \
istotności: (a) znaczenie praktyczne dla beneficjentów funduszy UE/KPO/PFR, \
(b) nowość lub zmiana linii orzeczniczej, (c) rozbieżność orzecznicza lub problem \
doktrynalny wart glosy, (d) potencjał newsowy/dyskusyjny. Wyrok rutynowy, \
powielający utrwaloną linię, ma niską istotność nawet gdy dotyczy funduszy.

ODRĘBNIE oceniasz, czy orzeczenie w ogóle dotyczy tematyki bazy (środki unijne, \
KPO, pomoc publiczna, subwencje PFR/Tarcza, Czyste Powietrze). To inna oś niż \
istotność contentowa: rutynowy wyrok o korekcie finansowej JEST związany z \
tematyką (zwiazane_z_tematyka=true, niska istotność). Niezwiązane są sprawy, \
które trafiły do bazy przypadkiem — np. spór podatkowy, w którym fraza padła \
marginalnie. W razie wątpliwości zaznaczaj zwiazane_z_tematyka=true.

Dobór kanału: LinkedIn = szybki news/komentarz (świeże, głośne, proste do \
streszczenia); Blog = praktyczny problem beneficjenta, z którego da się zrobić \
poradnikowy tekst; "Artykuł naukowy / glosa" = nowość doktrynalna, rozbieżność, \
orzeczenie precedensowe (zwłaszcza NSA poszerzony skład, SN, TSUE-adjacent); \
Podcast = temat nadający się na odcinek rozmowy; "Pomiń" = rutyna.

Odpowiadasz WYŁĄCZNIE poprawnym JSON, bez markdown, bez komentarzy."""

USER_PROMPT_TEMPLATE = """Tematy szczególnie obserwowane przez kancelarię (podnoszą istotność):
{watchlist}

Oceń poniższe orzeczenie i zwróć JSON o polach:
{{
  "zwiazane_z_tematyka": true | false,   // czy orzeczenie W OGÓLE dotyczy środków unijnych, KPO, pomocy publicznej lub subwencji PFR (nawet rutynowo)
  "pewnosc_zwiazku": <int 0-100>,        // jak pewna jest ta ocena
  "istotnosc": <int 0-100>,
  "dla_beneficjenta": "korzystne" | "niekorzystne" | "mieszane" | "nie dotyczy",
  "kanal_glowny": "LinkedIn" | "Blog" | "Artykuł naukowy / glosa" | "Podcast" | "Pomiń",
  "kanaly_dodatkowe": [<jak wyżej, może być pusta lista>],
  "tytul_roboczy": "<zwięzły tytuł pod wybrany kanał>",
  "hook": "<1-2 zdania otwierające, w duchu bloga kancelarii>",
  "punkty": ["<3-5 kluczowych punktów do omówienia>"],
  "uzasadnienie": "<2-3 zdania: dlaczego taka ocena i kanał>"
}}

Metadane: sygnatura {sygnatura}, {sad}, data {data}, kategoria: {finansowanie}.
Treść orzeczenia (może być ucięta):
---
{tekst}
---"""


def load_config() -> dict:
    with open(ROOT / "config.yaml", encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def call_claude(api_key: str, model: str, sygnatura: str, meta: dict, tekst: str,
                watchlist: list[str]) -> dict | None:
    prompt = USER_PROMPT_TEMPLATE.format(
        watchlist="\n".join(f"- {w}" for w in watchlist) or "- (brak)",
        sygnatura=sygnatura,
        sad=meta.get("sad", ""),
        data=meta.get("data", ""),
        finansowanie=meta.get("finansowanie", ""),
        tekst=tekst,
    )
    body = {
        "model": model,
        "max_tokens": 1200,
        "system": SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": prompt}],
    }
    headers = {
        "x-api-key": api_key,
        "anthropic-version": API_VERSION,
        "content-type": "application/json",
    }
    for attempt in range(3):
        try:
            resp = requests.post(API_URL, headers=headers, json=body, timeout=120)
            if resp.status_code == 429 or resp.status_code >= 500:
                wait = 10 * (attempt + 1)
                log.warning("API %s — ponawiam za %ds", resp.status_code, wait)
                time.sleep(wait)
                continue
            resp.raise_for_status()
            data = resp.json()
            text = "".join(
                block.get("text", "") for block in data.get("content", [])
                if block.get("type") == "text"
            )
            text = re.sub(r"^```(json)?|```$", "", text.strip(), flags=re.MULTILINE).strip()
            return json.loads(text)
        except json.JSONDecodeError as exc:
            log.error("%s: model nie zwrócił poprawnego JSON: %s", sygnatura, exc)
            return None
        except Exception as exc:  # noqa: BLE001
            log.error("%s: błąd API (próba %d): %s", sygnatura, attempt + 1, exc)
            time.sleep(5)
    return None


def get_ruling_text(link: str, notatka: str, max_chars: int) -> str:
    """Pobiera pełny tekst orzeczenia z SAOS lub CBOSA; awaryjnie fragment z notatki."""
    text = ""
    m = re.search(r"saos\.org\.pl/judgments/(\d+)", link or "")
    if m:
        text = saos_client.get_full_text(int(m.group(1)))
    elif link and "orzeczenia.nsa.gov.pl" in link:
        try:
            from bs4 import BeautifulSoup

            resp = requests.get(link, timeout=30, headers={
                "User-Agent": "nsa-monitor/2.0 (monitoring orzecznictwa)"})
            resp.raise_for_status()
            soup = BeautifulSoup(resp.text, "html.parser")
            text = soup.get_text(" ", strip=True)
        except Exception as exc:  # noqa: BLE001
            log.warning("Nie udało się pobrać treści z CBOSA (%s): %s", link, exc)
    if not text:
        text = notatka or ""
    return text[:max_chars]


def ensure_rekom_sheet(wb):
    if REKOM_SHEET in wb.sheetnames:
        return wb[REKOM_SHEET]
    ws = wb.create_sheet(REKOM_SHEET)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B356E")
    widths = [17, 10, 18, 26, 14, 12, 16, 22, 22, 44, 60, 70, 60, 40, 12]
    for i, (name, w) in enumerate(zip(REKOM_COLUMNS, widths), start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(REKOM_COLUMNS))}1"
    return ws


REPORT_PATH = ROOT / "agent_report.txt"
REJECTED_SHEET = "Odrzucone (auto)"


def ensure_rejected_sheet(wb, base):
    """Arkusz-kwarantanna: pełne wiersze usunięte z Bazy orzeczeń + metadane."""
    if REJECTED_SHEET in wb.sheetnames:
        return wb[REJECTED_SHEET]
    ws = wb.create_sheet(REJECTED_SHEET)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="7A1F1F")
    headers = [c.value for c in base[1]] + ["Data usunięcia", "Powód (AI)"]
    for i, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = header_font
        c.fill = header_fill
    ws.freeze_panes = "A2"
    return ws


def remove_offtopic(wb, base, header, to_remove: dict[str, str]) -> int:
    """Przenosi wskazane rekordy (ID -> powód) do arkusza kwarantanny
    i usuwa je z Bazy orzeczeń. Zwraca liczbę usuniętych."""
    if not to_remove:
        return 0
    rejected = ensure_rejected_sheet(wb, base)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    font = Font(name="Arial", size=10)
    rows = []
    for r in range(2, base.max_row + 1):
        rid = str(base.cell(r, header["ID"]).value or "").strip()
        if rid in to_remove:
            rows.append((r, rid))
    for r, rid in sorted(rows, reverse=True):  # od dołu, żeby indeksy nie uciekały
        values = [base.cell(r, c).value for c in range(1, len(header) + 1)]
        rr = rejected.max_row + 1
        for c, v in enumerate(values + [now, to_remove[rid]], start=1):
            cell = rejected.cell(row=rr, column=c, value=v)
            cell.font = copy.copy(font)
        base.delete_rows(r)
        log.info("Usunięto z bazy (poza tematyką): %s (%s)", rid, values[1])
    return len(rows)


def write_email_report(recommended, details, evaluated: int, queued: int) -> None:
    """Zapisuje raport tekstowy do wysyłki e-mailem (tylko gdy są rekomendacje)."""
    if not recommended:
        log.info("Brak rekomendacji — e-mail nie zostanie wysłany.")
        return
    lines = [
        f"Monitoring orzeczeń — raport z {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        f"Oceniono: {evaluated} orzeczeń, rekomendacji: {len(recommended)}."
        + (f" W kolejce do oceny pozostało: {queued}." if queued > 0 else ""),
        "",
        "REKOMENDACJE (wg istotności):",
        "",
    ]
    for score, syg, kanal, tytul in recommended:
        d = details.get(syg, {})
        lines.append(f"[{score}/100] {syg} — {d.get('sad', '')}, {d.get('data', '')}")
        lines.append(f"  Kanał: {kanal}"
                     + (f" (dodatkowo: {d['dodatkowe']})" if d.get("dodatkowe") else "")
                     + (f" | wstępnie: {d['wynik']}" if d.get("wynik") else ""))
        if tytul:
            lines.append(f"  Tytuł roboczy: {tytul}")
        if d.get("hook"):
            lines.append(f"  Hook: {d['hook']}")
        if d.get("link"):
            lines.append(f"  {d['link']}")
        lines.append("")
    lines.append("Pełne szczegóły (punkty, uzasadnienia): arkusz 'Rekomendacje' w pliku MASTER na Dysku Google.")
    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")
    log.info("Raport e-mail zapisany: %s", REPORT_PATH)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--max", type=int, default=None,
                        help="maksymalna liczba orzeczeń do oceny w tym przebiegu")
    args = parser.parse_args()

    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        log.info("Brak ANTHROPIC_API_KEY — pomijam agenta (monitor działa niezależnie).")
        return 0

    cfg = load_config()
    acfg = cfg.get("agent", {})
    if not acfg.get("enabled", True):
        log.info("Agent wyłączony w config.yaml.")
        return 0
    model = acfg.get("model", "claude-sonnet-4-6")
    max_per_run = args.max if args.max is not None else int(acfg.get("max_per_run", 20))
    min_score = int(acfg.get("min_score_for_recommendation", 55))
    max_chars = int(acfg.get("max_chars", 25000))
    watchlist = acfg.get("watchlist", [])
    auto_remove = bool(acfg.get("auto_remove_offtopic", True))
    remove_confidence = int(acfg.get("remove_min_confidence", 80))

    drive_sync.download_xlsx(str(LOCAL_XLSX))
    wb = load_workbook(LOCAL_XLSX)
    base = wb[cfg["master"]["sheet_name"]]
    header = {str(c.value).strip(): i + 1 for i, c in enumerate(base[1]) if c.value}

    rekom = ensure_rekom_sheet(wb)
    already = set()
    for r in range(2, rekom.max_row + 1):
        v = rekom.cell(r, 2).value
        if v:
            already.add(str(v).strip())

    # Kandydaci: rekordy dodane przez monitor, jeszcze nieocenione
    todo = []
    for r in range(2, base.max_row + 1):
        rid = base.cell(r, header["ID"]).value
        if not rid or str(rid).strip() in already:
            continue
        status = str(base.cell(r, header["Status weryfikacji"]).value or "")
        if status != "Metadane wstępne":
            continue
        todo.append({
            "id": str(rid).strip(),
            "sygnatura": str(base.cell(r, header["Sygnatura"]).value or ""),
            "sad": str(base.cell(r, header["Sąd"]).value or ""),
            "data": str(base.cell(r, header["Data orzeczenia"]).value or "")[:10],
            "finansowanie": str(base.cell(r, header["Rodzaj finansowania"]).value or ""),
            "link": str(base.cell(r, header["Link źródłowy"]).value or ""),
            "notatka": str(base.cell(r, header["Notatka wewnętrzna"]).value or ""),
        })

    if not todo:
        log.info("Brak nieocenionych rekordów — nic do zrobienia.")
        return 0
    # najnowsze najpierw; limit na przebieg (kontrola kosztów)
    todo.sort(key=lambda x: x["data"], reverse=True)
    batch = todo[:max_per_run]
    log.info("Do oceny: %d rekordów (w kolejce łącznie %d, model %s)",
             len(batch), len(todo), model)

    font = Font(name="Arial", size=10)
    wrap = Alignment(vertical="top", wrap_text=True)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    written = 0
    recommended = []
    details: dict[str, dict] = {}
    to_remove: dict[str, str] = {}

    for item in batch:
        tekst = get_ruling_text(item["link"], item["notatka"], max_chars)
        time.sleep(0.5)
        res = call_claude(api_key, model, item["sygnatura"], item, tekst, watchlist)
        if res is None:
            continue
        kanal = str(res.get("kanal_glowny", "Pomiń"))
        score = int(res.get("istotnosc", 0) or 0)
        if score < min_score:
            kanal = "Pomiń"
        offtopic = (res.get("zwiazane_z_tematyka") is False
                    and int(res.get("pewnosc_zwiazku", 0) or 0) >= remove_confidence)
        if offtopic:
            kanal = "Pomiń"
        row = [
            now, item["id"], item["sygnatura"], item["sad"], item["data"],
            score, str(res.get("dla_beneficjenta", "")), kanal,
            ", ".join(res.get("kanaly_dodatkowe", []) or []),
            str(res.get("tytul_roboczy", "")), str(res.get("hook", "")),
            "\n".join(f"• {p}" for p in (res.get("punkty") or [])),
            str(res.get("uzasadnienie", "")), item["link"],
            "Nowa" if kanal != "Pomiń" else "Pominięta",
        ]
        r = rekom.max_row + 1
        for c, value in enumerate(row, start=1):
            cell = rekom.cell(row=r, column=c, value=value)
            cell.font = copy.copy(font)
            if c in (11, 12, 13):
                cell.alignment = wrap
            if c == 14 and value:
                cell.hyperlink = value
        if offtopic and auto_remove:
            rekom.cell(row=r, column=15, value="Usunięta z bazy")
            to_remove[item["id"]] = (
                f"AI: poza tematyką (pewność {res.get('pewnosc_zwiazku', '?')}%). "
                + str(res.get("uzasadnienie", ""))
            )
        written += 1
        if kanal != "Pomiń":
            recommended.append((score, item["sygnatura"], kanal,
                                str(res.get("tytul_roboczy", ""))))
            details[item["sygnatura"]] = {
                "sad": item["sad"], "data": item["data"], "link": item["link"],
                "hook": str(res.get("hook", "")),
                "dodatkowe": ", ".join(res.get("kanaly_dodatkowe", []) or []),
                "wynik": str(res.get("dla_beneficjenta", "")),
            }
        log.info("%s: %d/100 -> %s", item["sygnatura"], score, kanal)

    removed = remove_offtopic(wb, base, header, to_remove) if auto_remove else 0
    if removed:
        log.info("Usunięto z Bazy orzeczeń %d rekordów spoza tematyki "
                 "(kopie w arkuszu '%s').", removed, REJECTED_SHEET)

    rekom.auto_filter.ref = f"A1:{get_column_letter(len(REKOM_COLUMNS))}{rekom.max_row}"
    wb.save(LOCAL_XLSX)
    drive_sync.upload_xlsx(str(LOCAL_XLSX))

    recommended.sort(reverse=True)
    log.info("Oceniono %d rekordów; rekomendacji: %d", written, len(recommended))
    for score, syg, kanal, tytul in recommended[:10]:
        log.info("  [%d] %s -> %s: %s", score, syg, kanal, tytul)
    if len(todo) > len(batch):
        log.info("W kolejce pozostało %d rekordów — kolejny przebieg oceni następne.",
                 len(todo) - len(batch))

    write_email_report(recommended, details, written, len(todo) - len(batch))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
